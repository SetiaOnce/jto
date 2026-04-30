from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, DataPenimbangan, DataSdm
from django.templatetags.static import static
from django.db.models import Q
from datetime import date, time, datetime, timedelta
from .utils import json_response
from .common import tablePersentaseKelebihan
import locale, os
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import calendar
from django.conf import settings

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Laporan Kelebihan',
        'active_menu': 'laporan_kelebihan',
        'app_name': lokasi.nama,
        'tablePersentaseKelebihan': tablePersentaseKelebihan(),
    }
    return render(request, 'backend/reports_kelebihan.html', context)


def show(request):
    if 'is_show_laporan' in request.GET:
        filter_interval = request.GET.get('filter_interval', 'MONTH').upper()
        today = date.today()
        current_month, current_year = today.month, today.year

        # Validasi bulan & tahun
        try:
            filter_month = int(request.GET.get('filter_month', current_month))
            if not 1 <= filter_month <= 12:
                raise ValueError
        except ValueError:
            return json_response(False, "Bulan tidak valid", 400)

        try:
            filter_year = int(request.GET.get('filter_year', current_year))
            if filter_year < 1:
                raise ValueError
        except ValueError:
            return json_response(False, "Tahun tidak valid", 400)

        # Buat list tanggal/bulan
        if filter_interval == 'MONTH':
            _, last_day = calendar.monthrange(filter_year, filter_month)
            date_list = [date(filter_year, filter_month, day) for day in range(1, last_day + 1)]
        elif filter_interval == 'YEAR':
            bulan_indo = [
                "Januari", "Februari", "Maret", "April", "Mei", "Juni",
                "Juli", "Agustus", "September", "Oktober", "November", "Desember"
            ]
            date_list = [{"bulan_id": m, "bulan_nama": bulan_indo[m - 1]} for m in range(1, 13)]
        else:
            return json_response(False, "filter_interval tidak valid", 400)

        # Helper: dapatkan range waktu dari d (naive datetime untuk USE_TZ = False)
        def get_range(d):
            if filter_interval == 'MONTH':
                return (
                    datetime.combine(d, time.min),
                    datetime.combine(d, time.max)
                )
            else:  # YEAR
                start = datetime(filter_year, d['bulan_id'], 1, 0, 0, 0)
                if d['bulan_id'] == 12:
                    end = datetime(filter_year + 1, 1, 1, 0, 0, 0)
                else:
                    end = datetime(filter_year, d['bulan_id'] + 1, 1, 0, 0, 0)
                return start, end

        data = []
        total_melanggar = 0
        total_under = 0
        total_kelebihan_per_jenis = [0] * len(tablePersentaseKelebihan()) 
        for d in date_list:
            query = DataPenimbangan.objects.filter(is_transaksi=1)

            if filter_interval == 'MONTH':
                is_selected = (d == today)
                time_name = d.strftime('%d-%b-%y')
            else:  # YEAR
                is_selected = (filter_year == current_year and d['bulan_id'] == current_month)
                time_name = d['bulan_nama']

            start_dt, end_dt = get_range(d)
            if filter_interval == 'MONTH':
                query = query.filter(tgl_penimbangan__range=(start_dt, end_dt))
            else:
                query = query.filter(tgl_penimbangan__gte=start_dt, tgl_penimbangan__lt=end_dt)

            kendaraan_melanggar = query.filter(is_melanggar='Y', prosen_lebih__gte=6).count()
            total_melanggar += kendaraan_melanggar

            kendaraan_under = query.filter(is_melanggar='Y', prosen_lebih__lte=5).count()
            total_under += kendaraan_under

            # Pelanggaran (dengan base query yang sudah di-filter)
            dt_kelebihan = []
            for idx, pl in enumerate(tablePersentaseKelebihan()):
                qs = query.all()  # ✅ Start dari query yang sudah di-filter, bukan .all()

                if pl['id'] == 2:
                    qs = qs.filter(prosen_lebih__gte=6, prosen_lebih__lte=20)
                elif pl['id'] == 3:
                    qs = qs.filter(prosen_lebih__gte=21, prosen_lebih__lte=40)
                elif pl['id'] == 4:
                    qs = qs.filter(prosen_lebih__gte=41, prosen_lebih__lte=60)
                elif pl['id'] == 5:
                    qs = qs.filter(prosen_lebih__gte=61, prosen_lebih__lte=80)
                elif pl['id'] == 6:
                    qs = qs.filter(prosen_lebih__gte=81, prosen_lebih__lte=100)
                elif pl['id'] == 7:
                    qs = qs.filter(prosen_lebih__gt=100)

                jumlah = qs.count()
                dt_kelebihan.append({'total': jumlah})
                total_kelebihan_per_jenis[idx] += jumlah

            data.append({
                'is_selected_column': is_selected,
                'time': time_name,
                'dt_kelebihan': dt_kelebihan,
                'kendaraan_melanggar': kendaraan_melanggar,
                'kendaraan_under': kendaraan_under,
            })
        data.append({
            'is_selected_column': True,
            "time": "TOTAL",
            'dt_kelebihan': [{'total': t} for t in total_kelebihan_per_jenis],
            'kendaraan_melanggar': total_melanggar,
            'kendaraan_under': total_under,
        })
    return json_response(status=True, message='Success', data=data)
    
def export(request):
    lokasi                      = LokasiUppkb.objects.first()
    filter_interval             = request.GET.get('filter_interval', 'MONTH').upper()
    today                       = date.today()
    current_month, current_year = today.month, today.year
    export_type                 = request.GET.get('export_type')

    # Validasi bulan & tahun
    try:
        filter_month = int(request.GET.get('filter_month', current_month))
        if not 1 <= filter_month <= 12:
            raise ValueError
    except ValueError:
        return json_response(False, "Bulan tidak valid", 400)

    try:
        filter_year = int(request.GET.get('filter_year', current_year))
        if filter_year < 1:
            raise ValueError
    except ValueError:
        return json_response(False, "Tahun tidak valid", 400)

    bulan_indo = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]
    # Buat list tanggal/bulan
    if filter_interval == 'MONTH':
        _, last_day = calendar.monthrange(filter_year, filter_month)
        date_list = [date(filter_year, filter_month, day) for day in range(1, last_day + 1)]
    elif filter_interval == 'YEAR':
        date_list = [{"bulan_id": m, "bulan_nama": bulan_indo[m - 1]} for m in range(1, 13)]
    else:
        return json_response(False, "filter_interval tidak valid", 400)

    # Helper: dapatkan range waktu dari d (naive datetime untuk USE_TZ = False)
    def get_range(d):
        if filter_interval == 'MONTH':
            return (
                datetime.combine(d, time.min),
                datetime.combine(d, time.max)
            )
        else:  # YEAR
            start = datetime(filter_year, d['bulan_id'], 1, 0, 0, 0)
            if d['bulan_id'] == 12:
                end = datetime(filter_year + 1, 1, 1, 0, 0, 0)
            else:
                end = datetime(filter_year, d['bulan_id'] + 1, 1, 0, 0, 0)
            return start, end

    data = []
    total_melanggar = 0
    total_under = 0
    total_kelebihan_per_jenis = [0] * len(tablePersentaseKelebihan()) 
    for d in date_list:
        query = DataPenimbangan.objects.filter(is_transaksi=1)

        if filter_interval == 'MONTH':
            is_selected = (d == today)
            time_name = d.strftime('%d-%b-%y')
        else:  # YEAR
            is_selected = (filter_year == current_year and d['bulan_id'] == current_month)
            time_name = d['bulan_nama']

        start_dt, end_dt = get_range(d)
        if filter_interval == 'MONTH':
            query = query.filter(tgl_penimbangan__range=(start_dt, end_dt))
        else:
            query = query.filter(tgl_penimbangan__gte=start_dt, tgl_penimbangan__lt=end_dt)

        kendaraan_melanggar = query.filter(is_melanggar='Y', prosen_lebih__gte=6).count()
        total_melanggar += kendaraan_melanggar

        kendaraan_under = query.filter(is_melanggar='Y', prosen_lebih__lte=5).count()
        total_under += kendaraan_under

        # Pelanggaran (dengan base query yang sudah di-filter)
        dt_kelebihan = []
        for idx, pl in enumerate(tablePersentaseKelebihan()):
            qs = query.all()  # ✅ Start dari query yang sudah di-filter, bukan .all()

            if pl['id'] == 2:
                qs = qs.filter(prosen_lebih__gte=6, prosen_lebih__lte=20)
            elif pl['id'] == 3:
                qs = qs.filter(prosen_lebih__gte=21, prosen_lebih__lte=40)
            elif pl['id'] == 4:
                qs = qs.filter(prosen_lebih__gte=41, prosen_lebih__lte=60)
            elif pl['id'] == 5:
                qs = qs.filter(prosen_lebih__gte=61, prosen_lebih__lte=80)
            elif pl['id'] == 6:
                qs = qs.filter(prosen_lebih__gte=81, prosen_lebih__lte=100)
            elif pl['id'] == 7:
                qs = qs.filter(prosen_lebih__gt=100)

            jumlah = qs.count()
            dt_kelebihan.append({'total': jumlah})
            total_kelebihan_per_jenis[idx] += jumlah

        data.append({
            'is_selected_column': False,
            'time': time_name,
            'dt_kelebihan': dt_kelebihan,
            'kendaraan_melanggar': kendaraan_melanggar,
            'kendaraan_under': kendaraan_under,
        })
    data.append({
        'is_selected_column': True,
        "time": "TOTAL",
        'dt_kelebihan': [{'total': t} for t in total_kelebihan_per_jenis],
        'kendaraan_melanggar': total_melanggar,
        'kendaraan_under': total_under,
    })
    if export_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Kelebihan Muatan"

        # ===== Header =====
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # Waktu
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # Total <= 5%
        ws.merge_cells(
            start_row=1,
            start_column=3,
            end_row=1,
            end_column=2 + len(tablePersentaseKelebihan())
        )  # % Kelebihan Muatan
        ws.merge_cells(
            start_row=1,
            start_column=3 + len(tablePersentaseKelebihan()),
            end_row=2,
            end_column=3 + len(tablePersentaseKelebihan())
        )  # Total Kendaraan Melanggar

        ws.cell(row=1, column=1).value = "WAKTU"
        ws.cell(row=1, column=2).value = "TOTAL <= 5% (TOLERANSI)"
        ws.cell(row=1, column=3).value = "% KELEBIHAN MUATAN (MELANGGAR)"
        ws.cell(row=1, column=3 + len(tablePersentaseKelebihan())).value = "TOTAL KENDARAAN MELANGGAR"

        # Header baris ke-2 untuk persentase muatan
        col = 3
        for pl in tablePersentaseKelebihan():
            ws.cell(row=2, column=col).value = pl['name']
            col += 1

        # ===== Data =====
        row_num = 3
        for row in data:
            ws.cell(row=row_num, column=1).value = row['time']
            ws.cell(row=row_num, column=2).value = row['kendaraan_under']

            col = 3
            for kl in row['dt_kelebihan']:
                ws.cell(row=row_num, column=col).value = kl['total']
                col += 1

            ws.cell(row=row_num, column=col).value = row['kendaraan_melanggar']

            row_num += 1

        # ===== Styling =====
        for r in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=col):
            for cell in r:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="BAD8B6", end_color="BAD8B6", fill_type="solid")

        for i in range(1, col + 1):
            ws.column_dimensions[chr(64 + i)].width = 18

        # ===== Kirim Response =====
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="Rekap_Kelebihan_Muatan_'
            + lokasi.nama.lower()
            + '_'
            + now().strftime("%Y%m%d%H%M%S")
            + '.xlsx"'
        )
        wb.save(response)
        return response
    elif export_type == 'pdf':
        nama_korsatpel = '-'
        nip_korsatpel = '-'
        pangkat_korsatpel = '-'
        ttd_korsatpel = None
        if lokasi.korsatpel_id:
            korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
            base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
            fallback = None
            def url_or_fallback(fn):
                if not fn:
                    return fallback
                if not os.path.isfile(os.path.join(base_dir, fn)):
                    return fallback
                return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

            if korsatpel:
                nama_korsatpel = korsatpel.name
                nip_korsatpel = korsatpel.nip
                pangkat_korsatpel = korsatpel.pangkat
                ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)

        context = {
            "lokasi": lokasi.nama.upper(),
            "tablePersentaseKelebihan": tablePersentaseKelebihan(),
            "periode_name": (
                bulan_indo[filter_month - 1].upper() + ' ' + str(filter_year)
                if filter_interval == 'MONTH'
                else str(filter_year)
            ),
            "korsatpel": {
                "nama": nama_korsatpel.upper(),
                "nip": nip_korsatpel.upper(),
                "pangkat": pangkat_korsatpel.upper(),
                "ttd": ttd_korsatpel,
            },
            "gambar": {
                "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
                "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
            },
            "data": data
        }
        html = render_to_string("exports/report_kelebihan.html", context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            'atachment; filename="Rekap_Kelebihan_Muatan_'
            + lokasi.nama.lower()
            + '_'
            + now().strftime("%Y%m%d%H%M%S")
            + '.pdf"'
        )

        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Terjadi kesalahan saat membuat PDF', status=500)
        return response