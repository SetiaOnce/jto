from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, KendaraanBodong, DataPelanggaran, MasterJenisPelanggaran, MasterKomoditi, DataKotaKab, MasterTimbangan, MasterRegu, MasterShift, MasterJenisKendaraan, UserSystem, DataSdm
from django.templatetags.static import static
from django.db.models import Q
from datetime import date, datetime, time
from .utils import json_response
import locale, os
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import random
from django.conf import settings

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Data Kendaraan Bodong',
        'active_menu': 'laporan_kendaraan_bodong',
        'app_name': lokasi.nama,
    }
    return render(request, 'backend/reports_kendaraan_bodong.html', context)

def show(request):
    try:
        # Params datatables
        draw = int(request.GET.get('draw', 0))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search = request.GET.get('search[value]', '').strip()

        filter_shift = request.GET.get('filter_shift')
        filter_regu = request.GET.get('filter_regu')
        filter_timbangan = request.GET.get('filter_timbangan')
        filter_date = str(request.GET.get('filter_date') or '')
        print("="*40)
        print("FILTER PARAMETER:")
        print(f"Shift     : {filter_shift or '-'}")
        print(f"Regu      : {filter_regu or '-'}")
        print(f"Timbangan : {filter_timbangan or '-'}")
        print(f"Tanggal   : {filter_date or '-'}")
        print("="*40)

        def parse_id_datetime(s: str, is_end: bool = False) -> datetime:
            """Parse datetime dari format Indonesian DD/MM/YYYY tanpa timezone"""
            s = (s or "").strip()
            for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    # kalau cuma tanggal, set jam default
                    if fmt == "%d/%m/%Y":
                        dt = datetime.combine(dt.date(), time.max if is_end else time.min)
                    return dt
                except ValueError:
                    continue
            raise ValueError(f"Format tanggal tidak valid: {s}")
        
        start_str, end_str = [x.strip() for x in filter_date.split(" - ")]

        start_datetime = parse_id_datetime(start_str, is_end=False)
        end_datetime   = parse_id_datetime(end_str, is_end=True)

        query = KendaraanBodong.objects.order_by("-tgl_penimbangan").filter(
            tgl_penimbangan__range=(start_datetime, end_datetime)
        )

        if filter_shift and filter_shift != 'ALL':
            query = query.filter(
                shift_id = filter_shift
            )
        if filter_regu and filter_regu != 'ALL':
            query = query.filter(
                regu_id = filter_regu
            )
        if filter_timbangan and filter_timbangan != 'ALL':
            query = query.filter(
                timbangan_id = filter_timbangan
            )
        if search:
            query = query.filter(
                Q(no_kendaraan__icontains=search) |
                Q(berat_timbang__icontains=search) |
                Q(panjang_ukur__icontains=search)|
                Q(lebar_ukur__icontains=search)|
                Q(tinggi_ukur__icontains=search)|
                Q(komoditi__icontains=search)|
                Q(asal_kota__icontains=search)|
                Q(tujuan_kota__icontains=search)
            )
        getRow = query.all()
        total_record = query.count()
        paginated_record = getRow[start:start+length]
        data = []

        for index, penimbangan in enumerate(paginated_record, start=start + 1):
            # Format waktu langsung tanpa timezone conversion
            tgl_penimbangan_formatted = penimbangan.tgl_penimbangan.strftime('%d/%m/%Y %H:%M')
            
            # Foto Penimbangan Depan
            foto_depan = static('dist/img/default-img.png')
            if penimbangan.foto_depan:
                if os.path.exists(os.path.join('static', 'dist', 'img', 'penimbangan', penimbangan.foto_depan)):
                    foto_depan = static('dist/img/penimbangan/') + penimbangan.foto_depan

            # Foto Penimbangan Belakang
            foto_belakang = static('dist/img/default-img.png')
            if penimbangan.foto_belakang:
                if os.path.exists(os.path.join('static', 'dist', 'img', 'penimbangan', penimbangan.foto_belakang)):
                    foto_belakang = static('dist/img/penimbangan/') + penimbangan.foto_belakang
            foto_depan = f'''
                <a class="image-popup" href="{foto_depan}" data-bs-toggle="tooltip" title="Klik untuk melihat gambar!">
                    <img src="{foto_depan}" class="rounded" alt="foto-penimbangan1" title="FOTO PENIMBANGAN - 1" width="72px">
                </a>
            '''
            foto_belakang = f'''
                <a class="image-popup" href="{foto_belakang}" data-bs-toggle="tooltip" title="Klik untuk melihat gambar!">
                    <img src="{foto_belakang}" class="rounded" alt="foto-penimbangan1" title="FOTO PENIMBANGAN - 2" width="72px">
                </a>
            '''
            # ==============================
            # KOMODITI, ASAL, TUJUAN, PENGADILAN
            # ==============================
            komoditi = MasterKomoditi.objects.filter(id=penimbangan.komoditi_id).select_related('kategori_komoditi').first()
            asal_kota = DataKotaKab.objects.filter(id=penimbangan.asal_kota_id).first()
            tujuan_kota = DataKotaKab.objects.filter(id=penimbangan.tujuan_kota_id).first()
            timbangan = MasterTimbangan.objects.filter(id=penimbangan.timbangan_id).first()
            regu = MasterRegu.objects.filter(id=penimbangan.regu_id).first()
            shift = MasterShift.objects.filter(id=penimbangan.shift_id).first()
            operator = UserSystem.objects.filter(id=penimbangan.petugas_id).first()
            
            data.append({
                'no': index,
                'foto_depan': foto_depan,
                'foto_belakang': foto_belakang,
                'waktu': tgl_penimbangan_formatted,
                'no_kendaraan': penimbangan.no_kendaraan,
                'berat_timbang': penimbangan.berat_timbang,
                'panjang_ukur': int(penimbangan.panjang_ukur),
                'lebar_ukur': int(penimbangan.lebar_ukur),
                'tinggi_ukur': int(penimbangan.tinggi_ukur),
                'komoditi': komoditi.nama if komoditi else '-',
                'asal_tujuan': (asal_kota.nama if asal_kota else '-') + ' - ' + (tujuan_kota.nama if tujuan_kota else '-'),
                'timbangan': timbangan.nama if timbangan else '-',
                'regu': regu.nama if regu else '-',
                'shift': shift.nama if shift else '-',
                'operator': operator.nama if operator else '-',
            })


        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_record,
            'recordsFiltered': getRow.count(),
            'data': data,
        })
    except Exception as e:
        return json_response(status=False, message=str(e), code=401) 
    
def export(request):
    try:
        export_type = request.GET.get('export_type')
        search = request.GET.get('search[value]', '')
        filter_date = request.GET.get('filter_date', '')
        filter_shift    = request.GET.get('filter_shift')
        filter_regu     = request.GET.get('filter_regu')
        filter_timbangan = request.GET.get('filter_timbangan')

        lokasi = LokasiUppkb.objects.first()

        # ============================
        # FILTER TANGGAL (NAIVE DATETIME)
        # ============================
        def parse_id_datetime(s: str, is_end: bool = False) -> datetime:
            """Parse datetime dari format Indonesian DD/MM/YYYY tanpa timezone"""
            s = (s or "").strip()
            for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    if fmt == "%d/%m/%Y":
                        dt = datetime.combine(dt.date(), time.max if is_end else time.min)
                    return dt
                except ValueError:
                    continue
            raise ValueError(f"Format tanggal tidak valid: {s}")
        
        if not filter_date:
            return json_response(status=False, message="Filter tanggal tidak boleh kosong", code=400)
        
        start_str, end_str = [x.strip() for x in filter_date.split(" - ")]
        start_datetime = parse_id_datetime(start_str, is_end=False)
        end_datetime   = parse_id_datetime(end_str, is_end=True)

        # ============================
        # QUERY
        # ============================
        query = KendaraanBodong.objects.order_by("-tgl_penimbangan").filter(
            tgl_penimbangan__range=(start_datetime, end_datetime)
        )
        if filter_shift and filter_shift != 'ALL':
            query = query.filter(shift_id = filter_shift)
        if filter_regu and filter_regu != 'ALL':
            query = query.filter(regu_id = filter_regu)
        if filter_timbangan and filter_timbangan != 'ALL':
            query = query.filter(timbangan_id = filter_timbangan)
        if search:
            query = query.filter(
                Q(no_kendaraan__icontains=search) |
                Q(berat_timbang__icontains=search) |
                Q(panjang_ukur__icontains=search)|
                Q(lebar_ukur__icontains=search)|
                Q(tinggi_ukur__icontains=search)|
                Q(komoditi__icontains=search)|
                Q(asal_kota__icontains=search)|
                Q(tujuan_kota__icontains=search)
            )

        # ============================
        # SERIALIZE DATA UNTUK EXCEL/PDF
        # ============================
        data = []
        for penimbangan in query:
            # Format waktu langsung tanpa timezone conversion
            tgl_penimbangan_formatted = penimbangan.tgl_penimbangan.strftime('%d/%m/%Y %H:%M')
            
            # KOMODITI, ASAL, TUJUAN
            komoditi = MasterKomoditi.objects.filter(id=penimbangan.komoditi_id).select_related('kategori_komoditi').first()
            asal_kota = DataKotaKab.objects.filter(id=penimbangan.asal_kota_id).first()
            tujuan_kota = DataKotaKab.objects.filter(id=penimbangan.tujuan_kota_id).first()
            timbangan = MasterTimbangan.objects.filter(id=penimbangan.timbangan_id).first()
            regu = MasterRegu.objects.filter(id=penimbangan.regu_id).first()
            shift = MasterShift.objects.filter(id=penimbangan.shift_id).first()
            operator = UserSystem.objects.filter(id=penimbangan.petugas_id).first()

            data.append({
                'waktu': tgl_penimbangan_formatted,
                'no_kendaraan': penimbangan.no_kendaraan,
                'berat_timbang': penimbangan.berat_timbang,
                'panjang_ukur': int(penimbangan.panjang_ukur),
                'lebar_ukur': int(penimbangan.lebar_ukur),
                'tinggi_ukur': int(penimbangan.tinggi_ukur),
                'komoditi': komoditi.nama if komoditi else '-',
                'asal_tujuan': (asal_kota.nama if asal_kota else '-') + ' - ' + (tujuan_kota.nama if tujuan_kota else '-'),
                'timbangan': timbangan.nama if timbangan else '-',
                'regu': regu.nama if regu else '-',
                'shift': shift.nama if shift else '-',
                'operator': operator.nama if operator else '-',
            })

        if not data:
            return json_response(status=False, message="Tidak ada data untuk di-export", code=400)
    
        # ============================
        # EXPORT EXCEL
        # ============================
        if export_type == "excel":
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Kendaraan Bodong"

                HEADER_MAP = {
                    "waktu": "WAKTU PENIMBANGAN",
                    "no_kendaraan": "NO KENDARAAN",
                    "komoditi": "KOMODITI",
                    "asal_tujuan": "ASAL -TUJUAN",
                    "berat_timbang": "BERAT(KG)",
                    "panjang_ukur": "PANJANG UKUR(MM)",
                    "lebar_ukur": "LEBAR UKUR(MM)",
                    "tinggi_ukur": "TINGGI UKUR(MM)",
                    "timbangan": "TIMBANGAN",
                    "operator": "OPERATOR",
                    "shift": "SHIFT",
                    "regu": "REGU",
                }
                headers = [HEADER_MAP.get(k, k.upper()) for k in data[0].keys()] if data else []
                ws.append(headers)

                # ===== Styling =====
                orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

                # Border tebal untuk header (medium)
                header_border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium')
                )

                # Border tipis untuk body data
                body_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # ===== SET HEADER =====
                for h in ws[1]:
                    h.font = Font(bold=True, color="FFFFFF")
                    h.fill = orange_fill
                    h.alignment = Alignment(horizontal="center", vertical="center")
                    h.border = header_border

                # ===== SET BODY =====
                # Append data rows
                for row in data:
                    ws.append(list(row.values()))

                # Apply border ke setiap cell di body
                row_start = 2
                row_end = ws.max_row
                col_start = 1
                col_end = ws.max_column

                for r in range(row_start, row_end + 1):
                    for c in range(col_start, col_end + 1):
                        ws.cell(row=r, column=c).border = body_border

                for col in ws.columns:
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws.column_dimensions[col[0].column_letter].width = max_len + 2

                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                timestamp = now().strftime("%Y_%m_%d_%H_%M_%S_%f")
                random_code = random.randint(10000000, 99999999)
                filename = f'kendaraan_bodong_{timestamp}_{random_code}.xlsx'
                response["Content-Disposition"] = f'attachment; filename="{filename}"'

                wb.save(response)
                return response
            except Exception as e:
                return json_response(status=False, message=f"Error export Excel: {str(e)}", code=500)

        # ============================
        # EXPORT PDF
        # ============================
        elif export_type == "pdf":
            try:
                nama_korsatpel = '-'
                nip_korsatpel = '-'
                pangkat_korsatpel = '-'
                ttd_korsatpel = None
                if lokasi.korsatpel_id:
                    korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
                    base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
                    fallback = None
                    def url_or_fallback(fn):
                        if not fn:
                            return fallback
                        if not os.path.isfile(os.path.join(base_dir, fn)):
                            return fallback
                        return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

                    if korsatpel:
                        nama_korsatpel = korsatpel.name
                        nip_korsatpel = korsatpel.nip
                        pangkat_korsatpel = korsatpel.pangkat
                        ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)
                        
                context = {
                    "lokasi": lokasi.nama.upper(),
                    "periode_name": filter_date,
                    "date_now": now().strftime("%d-%m-%Y"),
                    "shift": MasterShift.objects.get(id=filter_shift).nama if filter_shift else "SEMUA",
                    "regu": MasterRegu.objects.get(id=filter_regu).nama if filter_regu else "SEMUA",
                    "korsatpel": {
                        "nama": nama_korsatpel.upper(),
                        "nip": nip_korsatpel.upper(),
                        "pangkat": pangkat_korsatpel.upper(),
                        "ttd": ttd_korsatpel,
                    },
                    "gambar": {
                        "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
                        "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
                    },
                    "data": data,
                }

                html = render_to_string("exports/report_kendaraan_bodong.html", context)

                response = HttpResponse(content_type="application/pdf")
                response["Content-Disposition"] = (
                    f'attachment; filename="data_kendaraan_bodong_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
                )

                pisa_status = pisa.CreatePDF(html, dest=response)

                if pisa_status.err:
                    return json_response(status=False, message="Gagal membuat PDF", code=500)

                return response
            except Exception as e:
                return json_response(status=False, message=f"Error export PDF: {str(e)}", code=500)
        
        # Jika export_type tidak valid
        return json_response(status=False, message="Export type tidak valid", code=400)
    
    except Exception as e:
        return json_response(status=False, message=str(e), code=500)