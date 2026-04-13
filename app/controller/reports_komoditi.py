from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, DataPenimbangan, MasterKomoditi, DataSdm, MasterKomoditi
from django.templatetags.static import static
from django.db.models import Q
from datetime import date
from .utils import json_response, convert_timezone
import locale, os
from datetime import datetime, date, time
from django.utils.timezone import now, make_aware, get_current_timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import random
from django.conf import settings
import calendar
from django.utils import timezone
from django.db.models import Count

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Data Komoditi',
        'active_menu': 'laporan_komoditi',
        'app_name': lokasi.nama,
    }
    return render(request, 'backend/reports_komoditi.html', context)


def show(request):
    if 'is_show_laporan' in request.GET:
        tz = timezone.get_current_timezone()
        filter_date = request.GET.get('filter_date')

        if filter_date:
            start_str, end_str = filter_date.split(" - ")

            start_dt = timezone.make_aware(
                datetime.strptime(start_str, "%Y-%m-%d %H:%M"),
                tz
            )
            end_dt = timezone.make_aware(
                datetime.strptime(end_str, "%Y-%m-%d %H:%M"),
                tz
            )
        else:
            # fallback hari ini
            today = timezone.now()
            start_dt = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end_dt   = today.replace(hour=23, minute=59, second=59, microsecond=999999)

        # ===== Query utama =====
        summary = (
            DataPenimbangan.objects
            .filter(is_transaksi=1, tgl_penimbangan__range=(start_dt, end_dt))
            .values('komoditi_id')
            .annotate(total=Count('id'))
        )

        # jadikan dict: {komoditi_id: total}
        summary_map = {row['komoditi_id']: row['total'] for row in summary}

        # ===== 2. Ambil hanya MasterKomoditi yang ADA di transaksi
        komoditi_ids = summary_map.keys()

        komoditi_qs = (
            MasterKomoditi.objects
            .filter(id__in=komoditi_ids, is_active=1)
            .select_related('kategori_komoditi')
            .order_by('kategori_komoditi__id', 'id')
        )

        # ===== 3. Susun group per kategori
        groups_map = {}
        grand_total = 0

        for k in komoditi_qs:
            kategori = k.kategori_komoditi

            if kategori.id not in groups_map:
                groups_map[kategori.id] = {
                    "code": kategori.kode,
                    "title": kategori.nama,
                    "items": [],
                    "total": 0
                }

            total = summary_map.get(k.id, 0)
            grand_total += total

            groups_map[kategori.id]["items"].append({
                "no": len(groups_map[kategori.id]["items"]) + 1,
                "key": str(k.id),
                "name": k.nama,
                "total": total
            })

            groups_map[kategori.id]["total"] += total

        # ===== 4. Output final
        groups = list(groups_map.values())

        payload = {
            "groups": groups,
            "grand_total": grand_total
        }

        return json_response(True, "Success", 200, payload)

def export(request):
    lokasi = LokasiUppkb.objects.first()
    tz = timezone.get_current_timezone()

    export_type = request.GET.get('export_type')
    filter_date = request.GET.get('filter_date')

    if filter_date:
        start_str, end_str = filter_date.split(" - ")

        start_dt = timezone.make_aware(
            datetime.strptime(start_str, "%Y-%m-%d %H:%M"),
            tz
        )
        end_dt = timezone.make_aware(
            datetime.strptime(end_str, "%Y-%m-%d %H:%M"),
            tz
        )
    else:
        # fallback hari ini
        today = timezone.now()
        start_dt = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt   = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    # ===== Query utama =====
    summary = (
        DataPenimbangan.objects
        .filter(is_transaksi=1, tgl_penimbangan__range=(start_dt, end_dt))
        .values('komoditi_id')
        .annotate(total=Count('id'))
    )

    # jadikan dict: {komoditi_id: total}
    summary_map = {row['komoditi_id']: row['total'] for row in summary}

    # ===== 2. Ambil hanya MasterKomoditi yang ADA di transaksi
    komoditi_ids = summary_map.keys()

    komoditi_qs = (
        MasterKomoditi.objects
        .filter(id__in=komoditi_ids, is_active=1)
        .select_related('kategori_komoditi')
        .order_by('kategori_komoditi__id', 'id')
    )

    # ===== 3. Susun group per kategori
    groups_map = {}
    grand_total = 0

    for k in komoditi_qs:
        kategori = k.kategori_komoditi

        if kategori.id not in groups_map:
            groups_map[kategori.id] = {
                "code": kategori.kode,
                "title": kategori.nama,
                "items": [],
                "total": 0
            }

        total = summary_map.get(k.id, 0)
        grand_total += total

        groups_map[kategori.id]["items"].append({
            "no": len(groups_map[kategori.id]["items"]) + 1,
            "key": str(k.id),
            "name": k.nama,
            "total": total
        })

        groups_map[kategori.id]["total"] += total

    # ===== 4. Output final
    groups = list(groups_map.values())
    payload = {
        "groups": groups,
        "grand_total": grand_total
    }

    if export_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Komoditi"

        # =========================
        # STYLING
        # =========================
        orange_fill = PatternFill(start_color="FF7B1A", end_color="FF7B1A", fill_type="solid")
        group_fill  = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        thin = Side(style="thin", color="CFCFCF")
        body_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_font = Font(bold=True, color="FFFFFF")
        bold_font   = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # =========================
        # HEADER / INFO
        # =========================
        # Baris periode (optional)
        ws.merge_cells("A1:D1")
        ws["A1"] = f"PERIODE: {filter_date or (start_dt.strftime('%Y-%m-%d %H:%M') + ' - ' + end_dt.strftime('%Y-%m-%d %H:%M'))}"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = left

        # Header tabel
        header_row = 3
        headers = ["KEL", "NO", "KOMODITI", "TOTAL"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = body_border

        # Set lebar kolom (stabil)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 12

        # =========================
        # BODY (groups/items/totals)
        # =========================
        row = header_row + 1  # mulai isi setelah header

        for g in payload["groups"]:
            # --- Group Title (merge A-D)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = ws.cell(row=row, column=1, value=str(g["title"]).upper())
            c.fill = group_fill
            c.font = bold_font
            c.alignment = left
            # border untuk merge area
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = body_border
                ws.cell(row=row, column=col).fill = group_fill
            row += 1

            items = g.get("items", [])
            start_item_row = row

            # --- Item rows
            for it in items:
                ws.cell(row=row, column=1, value=g["code"]).alignment = center
                ws.cell(row=row, column=2, value=it["no"]).alignment = center
                ws.cell(row=row, column=3, value=str(it["name"]).upper()).alignment = left
                ws.cell(row=row, column=4, value=it["total"]).alignment = center

                # apply border
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = body_border

                row += 1

            # --- Merge KEL (kolom A) agar seperti rowspan (opsional tapi mirip template kamu)
            if items:
                ws.merge_cells(start_row=start_item_row, start_column=1, end_row=row-1, end_column=1)
                # pastikan align center setelah merge
                ws.cell(row=start_item_row, column=1).alignment = center
                ws.cell(row=start_item_row, column=1).font = bold_font

            # --- Total per group (merge A-C)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value=f"TOTAL {g['title']}").alignment = center
            ws.cell(row=row, column=1).font = bold_font

            ws.cell(row=row, column=4, value=g.get("total", 0)).alignment = center
            ws.cell(row=row, column=4).font = bold_font

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = body_border

            row += 2  # 1 baris kosong antar group

        # =========================
        # GRAND TOTAL
        # =========================
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="GRAND TOTAL").alignment = center
        ws.cell(row=row, column=1).font = Font(bold=True)

        ws.cell(row=row, column=4, value=payload.get("grand_total", 0)).alignment = center
        ws.cell(row=row, column=4).font = Font(bold=True)

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = body_border

        # =========================
        # RESPONSE
        # =========================
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = now().strftime("%Y%m%d_%H%M%S")
        random_code = random.randint(100000, 999999)
        filename = f"rekap_komoditi_{lokasi.nama.lower()}_{timestamp}_{random_code}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
    
    # ============================
    # EXPORT PDF
    # ============================
    nama_korsatpel = '-'
    nip_korsatpel = '-'
    pangkat_korsatpel = '-'
    ttd_korsatpel = None
    if lokasi.korsatpel_id:
        korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
        base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
        fallback = None
        def url_or_fallback(fn):
            if not fn:
                return fallback
            if not os.path.isfile(os.path.join(base_dir, fn)):
                return fallback
            return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

        if korsatpel:
            nama_korsatpel = korsatpel.name
            nip_korsatpel = korsatpel.nip
            pangkat_korsatpel = korsatpel.pangkat
            ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)
            
    context = {
        "lokasi": lokasi.nama.upper(),
        "periode_name": filter_date,
        "date_now": now().strftime("%d-%m-%Y"),
        "korsatpel": {
            "nama": nama_korsatpel.upper(),
            "nip": nip_korsatpel.upper(),
            "pangkat": pangkat_korsatpel.upper(),
            "ttd": ttd_korsatpel,
        },
        "gambar": {
            "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
            "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
        },
        "data": payload,
    }

    html = render_to_string("exports/report_komoditi.html", context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        # f'inline; filename="data_komoditi_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
        f'atachment; filename="data_komoditi_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)

    return response