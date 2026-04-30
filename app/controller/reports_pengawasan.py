from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, DataPenimbangan, DataSdm, DataPelanggaran, DataPenindakan
from django.templatetags.static import static
from django.db.models import Q
from datetime import date, time, datetime, timedelta
from .utils import json_response
from .common import jenisPelanggaran, sanksi, subSanksi
import locale, os
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import calendar
from django.conf import settings

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Laporan Pengawasan',
        'active_menu': 'laporan',
        'app_name': lokasi.nama,
        'jenisPelanggaran': jenisPelanggaran(),
        'sanksi': sanksi(),
        'subSanksi': subSanksi(),
    }
    return render(request, 'backend/reports_pengawasan.html', context)


def show(request):
    if 'is_show_laporan' in request.GET:
        filter_interval = request.GET.get('filter_interval', 'MONTH').upper()
        today = date.today()
        current_month, current_year = today.month, today.year

        # Validasi bulan & tahun
        try:
            filter_month = int(request.GET.get('filter_month', current_month))
            if not 1 <= filter_month <= 12:
                raise ValueError
        except ValueError:
            return json_response(False, "Bulan tidak valid", 400)

        try:
            filter_year = int(request.GET.get('filter_year', current_year))
            if filter_year < 1:
                raise ValueError
        except ValueError:
            return json_response(False, "Tahun tidak valid", 400)

        # Buat list tanggal/bulan
        if filter_interval == 'MONTH':
            _, last_day = calendar.monthrange(filter_year, filter_month)
            date_list = [date(filter_year, filter_month, day) for day in range(1, last_day + 1)]
        elif filter_interval == 'YEAR':
            bulan_indo = [
                "Januari", "Februari", "Maret", "April", "Mei", "Juni",
                "Juli", "Agustus", "September", "Oktober", "November", "Desember"
            ]
            date_list = [{"bulan_id": m, "bulan_nama": bulan_indo[m - 1]} for m in range(1, 13)]
        else:
            return json_response(False, "filter_interval tidak valid", 400)

        # Helper: dapatkan range waktu dari d (naive datetime untuk USE_TZ = False)
        def get_range(d):
            if filter_interval == 'MONTH':
                return (
                    datetime.combine(d, time.min),
                    datetime.combine(d, time.max)
                )
            else:  # YEAR
                start = datetime(filter_year, d['bulan_id'], 1, 0, 0, 0)
                if d['bulan_id'] == 12:
                    end = datetime(filter_year + 1, 1, 1, 0, 0, 0)
                else:
                    end = datetime(filter_year, d['bulan_id'] + 1, 1, 0, 0, 0)
                return start, end

        data = []
        total_diperiksa = 0
        total_melanggar = 0
        total_tidak_melanggar = 0
        total_pelanggaran_per_jenis = [0] * len(jenisPelanggaran()) 
        total_sanksi_per_jenis = [0] * len(sanksi()) 
        total_subsanksi_per_jenis = [0] * len(subSanksi()) 
        for d in date_list:
            jumlah_kendaraan = DataPenimbangan.objects.filter(is_transaksi=1)

            if filter_interval == 'MONTH':
                is_selected = (d == today)
                time_name = d
            else:  # YEAR
                is_selected = (filter_year == current_year and d['bulan_id'] == current_month)
                time_name = d['bulan_nama']

            start_dt, end_dt = get_range(d)
            if filter_interval == 'MONTH':
                jumlah_kendaraan = jumlah_kendaraan.filter(tgl_penimbangan__range=(start_dt, end_dt))
            else:
                jumlah_kendaraan = jumlah_kendaraan.filter(tgl_penimbangan__gte=start_dt, tgl_penimbangan__lt=end_dt)

            diperiksa = jumlah_kendaraan.count()
            melanggar = jumlah_kendaraan.filter(is_melanggar='Y').count()
            tidak_melanggar = jumlah_kendaraan.filter(is_melanggar='N').count()
            # --- akumulasi total ---
            total_diperiksa += diperiksa
            total_melanggar += melanggar
            total_tidak_melanggar += tidak_melanggar

            # Pelanggaran
            dt_pelanggaran = []
            for idx, pl in enumerate(jenisPelanggaran()):
                qs = DataPelanggaran.objects.filter(jenis_pelanggaran_id=pl['id'])
                if filter_interval == 'MONTH':
                    qs = qs.filter(tgl_penimbangan=d)
                else:
                    qs = qs.filter(tgl_penimbangan__year=filter_year, tgl_penimbangan__month=d['bulan_id'])
                jumlah = qs.count()
                dt_pelanggaran.append({'total': jumlah})
                total_pelanggaran_per_jenis[idx] += jumlah

            # Penindakan
            dt_penindakan = []
            for idx, sn in enumerate(sanksi()):
                qs = DataPenindakan.objects.filter(sanksi_id=sn['id'])
                qs = qs.filter(tgl_penindakan__gte=start_dt, tgl_penindakan__lt=end_dt) if filter_interval == 'YEAR' \
                    else qs.filter(tgl_penindakan__range=(start_dt, end_dt))
                jumlah = qs.count()
                dt_penindakan.append({'total': jumlah})
                total_sanksi_per_jenis[idx] += jumlah
            # Sub Sanksi
            dt_subsanksi = []
            for idx, ss in enumerate(subSanksi()):
                target_id = str(ss['id'])
                qs = DataPenindakan.objects.filter(
                    Q(sanksi_tambahan=target_id) |
                    Q(sanksi_tambahan__startswith=target_id + ',') |
                    Q(sanksi_tambahan__endswith=',' + target_id) |
                    Q(sanksi_tambahan__contains=',' + target_id + ',')
                )
                qs = qs.filter(tgl_penindakan__gte=start_dt, tgl_penindakan__lt=end_dt) if filter_interval == 'YEAR' \
                    else qs.filter(tgl_penindakan__range=(start_dt, end_dt))
                jumlah = qs.count()
                dt_subsanksi.append({'total': jumlah})
                total_subsanksi_per_jenis[idx] += jumlah

            data.append({
                'is_selected_column': is_selected,
                'time': time_name,
                'diperiksa': diperiksa,
                'melanggar': melanggar,
                'tidak_melanggar': tidak_melanggar,
                'dt_pelanggaran': dt_pelanggaran,
                'dt_penindakan': dt_penindakan,
                'dt_subsanksi': dt_subsanksi,
            })
        data.append({
            'is_selected_column': True,
            "time": "TOTAL",
            "diperiksa": total_diperiksa,
            "melanggar": total_melanggar,
            "tidak_melanggar": total_tidak_melanggar,
            'dt_pelanggaran': [{'total': t} for t in total_pelanggaran_per_jenis],
            'dt_penindakan': [{'total': t} for t in total_sanksi_per_jenis],
            'dt_subsanksi': [{'total': t} for t in total_subsanksi_per_jenis],
        })
    return json_response(status=True, message='Success', data=data)
    
def export(request):
    lokasi                      = LokasiUppkb.objects.first()
    filter_interval = request.GET.get('filter_interval', 'MONTH').upper()
    today = date.today()
    current_month, current_year = today.month, today.year
    export_type = request.GET.get('export_type')

    # Validasi bulan & tahun
    try:
        filter_month = int(request.GET.get('filter_month', current_month))
        if not 1 <= filter_month <= 12:
            raise ValueError
    except ValueError:
        return json_response(False, "Bulan tidak valid", 400)

    try:
        filter_year = int(request.GET.get('filter_year', current_year))
        if filter_year < 1:
            raise ValueError
    except ValueError:
        return json_response(False, "Tahun tidak valid", 400)

    bulan_indo = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]
    # Buat list tanggal/bulan
    if filter_interval == 'MONTH':
        _, last_day = calendar.monthrange(filter_year, filter_month)
        date_list = [date(filter_year, filter_month, day) for day in range(1, last_day + 1)]
    elif filter_interval == 'YEAR':
        date_list = [{"bulan_id": m, "bulan_nama": bulan_indo[m - 1]} for m in range(1, 13)]
    else:
        return json_response(False, "filter_interval tidak valid", 400)

    # Helper: dapatkan range waktu dari d (naive datetime untuk USE_TZ = False)
    def get_range(d):
        if filter_interval == 'MONTH':
            return (
                datetime.combine(d, time.min),
                datetime.combine(d, time.max)
            )
        else:  # YEAR
            start = datetime(filter_year, d['bulan_id'], 1, 0, 0, 0)
            if d['bulan_id'] == 12:
                end = datetime(filter_year + 1, 1, 1, 0, 0, 0)
            else:
                end = datetime(filter_year, d['bulan_id'] + 1, 1, 0, 0, 0)
            return start, end

    data = []
    total_diperiksa = 0
    total_melanggar = 0
    total_tidak_melanggar = 0
    total_pelanggaran_per_jenis = [0] * len(jenisPelanggaran()) 
    total_sanksi_per_jenis = [0] * len(sanksi()) 
    total_subsanksi_per_jenis = [0] * len(subSanksi()) 
    for d in date_list:
        jumlah_kendaraan = DataPenimbangan.objects.filter(is_transaksi=1)

        if filter_interval == 'MONTH':
            is_selected = (d == today)
            time_name = d.strftime("%Y-%m-%d")
        else:  # YEAR
            is_selected = (filter_year == current_year and d['bulan_id'] == current_month)
            time_name = d['bulan_nama']

        start_dt, end_dt = get_range(d)
        if filter_interval == 'MONTH':
            jumlah_kendaraan = jumlah_kendaraan.filter(tgl_penimbangan__range=(start_dt, end_dt))
        else:
            jumlah_kendaraan = jumlah_kendaraan.filter(tgl_penimbangan__gte=start_dt, tgl_penimbangan__lt=end_dt)

        diperiksa = jumlah_kendaraan.count()
        melanggar = jumlah_kendaraan.filter(is_melanggar='Y').count()
        tidak_melanggar = jumlah_kendaraan.filter(is_melanggar='N').count()
        # --- akumulasi total ---
        total_diperiksa += diperiksa
        total_melanggar += melanggar
        total_tidak_melanggar += tidak_melanggar

        # Pelanggaran
        dt_pelanggaran = []
        for idx, pl in enumerate(jenisPelanggaran()):
            qs = DataPelanggaran.objects.filter(jenis_pelanggaran_id=pl['id'])
            if filter_interval == 'MONTH':
                qs = qs.filter(tgl_penimbangan=d)
            else:
                qs = qs.filter(tgl_penimbangan__year=filter_year, tgl_penimbangan__month=d['bulan_id'])
            jumlah = qs.count()
            dt_pelanggaran.append({'total': jumlah})
            total_pelanggaran_per_jenis[idx] += jumlah

        # Penindakan
        dt_penindakan = []
        for idx, sn in enumerate(sanksi()):
            qs = DataPenindakan.objects.filter(sanksi_id=sn['id'])
            qs = qs.filter(tgl_penindakan__gte=start_dt, tgl_penindakan__lt=end_dt) if filter_interval == 'YEAR' \
                else qs.filter(tgl_penindakan__range=(start_dt, end_dt))
            jumlah = qs.count()
            dt_penindakan.append({'total': jumlah})
            total_sanksi_per_jenis[idx] += jumlah
        # Sub Sanksi
        dt_subsanksi = []
        for idx, ss in enumerate(subSanksi()):
            target_id = str(ss['id'])
            qs = DataPenindakan.objects.filter(
                Q(sanksi_tambahan=target_id) |
                Q(sanksi_tambahan__startswith=target_id + ',') |
                Q(sanksi_tambahan__endswith=',' + target_id) |
                Q(sanksi_tambahan__contains=',' + target_id + ',')
            )
            qs = qs.filter(tgl_penindakan__gte=start_dt, tgl_penindakan__lt=end_dt) if filter_interval == 'YEAR' \
                else qs.filter(tgl_penindakan__range=(start_dt, end_dt))
            jumlah = qs.count()
            dt_subsanksi.append({'total': jumlah})
            total_subsanksi_per_jenis[idx] += jumlah

        data.append({
            'is_selected_column': False,
            'time': time_name,
            'diperiksa': diperiksa,
            'melanggar': melanggar,
            'tidak_melanggar': tidak_melanggar,
            'dt_pelanggaran': dt_pelanggaran,
            'dt_penindakan': dt_penindakan,
            'dt_subsanksi': dt_subsanksi,
        })

    data.append({
        'is_selected_column': True,
        "time": "TOTAL",
        "diperiksa": total_diperiksa,
        "melanggar": total_melanggar,
        "tidak_melanggar": total_tidak_melanggar,
        'dt_pelanggaran': [{'total': t} for t in total_pelanggaran_per_jenis],
        'dt_penindakan': [{'total': t} for t in total_sanksi_per_jenis],
        'dt_subsanksi': [{'total': t} for t in total_subsanksi_per_jenis],
    })
    if export_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pengawasan"

        # ====== Buat Header Baris 1 ======
        headers_row1 = [
            "WAKTU",  # kolom pertama
            "JUMLAH KENDARAAN",  # merge 3 kolom di bawah
            "PELANGGARAN",
            "PENINDAKAN",
            "PENINDAKAN LAINNYA"
        ]
        # Dinamis: hitung kolom masing-masing bagian
        jumlah_pelanggaran = len(jenisPelanggaran())
        jumlah_sanksi = len(sanksi())
        jumlah_subsanksi = len(subSanksi())

        # Tulis header level 1
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # Waktu merge vertikal
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)  # Jumlah Kendaraan
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=4 + jumlah_pelanggaran)  # Pelanggaran
        ws.merge_cells(
            start_row=1,
            start_column=5 + jumlah_pelanggaran,
            end_row=1,
            end_column=4 + jumlah_pelanggaran + jumlah_sanksi
        )
        ws.merge_cells(
            start_row=1,
            start_column=5 + jumlah_pelanggaran + jumlah_sanksi,
            end_row=1,
            end_column=4 + jumlah_pelanggaran + jumlah_sanksi + jumlah_subsanksi
        )

        ws.cell(row=1, column=1).value = "WAKTU"
        ws.cell(row=1, column=2).value = "JUMLAH KENDARAAN"
        ws.cell(row=1, column=5).value = "PELANGGARAN"
        ws.cell(row=1, column=5 + jumlah_pelanggaran).value = "PENINDAKAN"
        ws.cell(row=1, column=5 + jumlah_pelanggaran + jumlah_sanksi).value = "PENINDAKAN LAINNYA"

        # ====== Buat Header Baris 2 ======
        ws.cell(row=2, column=2).value = "DIPERIKSA"
        ws.cell(row=2, column=3).value = "MELANGGAR"
        ws.cell(row=2, column=4).value = "TIDAK MELANGGAR"

        col = 5
        for pl in jenisPelanggaran():
            ws.cell(row=2, column=col).value = pl['name']
            col += 1
        for sn in sanksi():
            ws.cell(row=2, column=col).value = sn['name']
            col += 1
        for ss in subSanksi():
            ws.cell(row=2, column=col).value = ss['name']
            col += 1

        # ====== Tulis Data ======
        row_num = 3
        for row in data:
            ws.cell(row=row_num, column=1).value = row['time']
            ws.cell(row=row_num, column=2).value = row['diperiksa']
            ws.cell(row=row_num, column=3).value = row['melanggar']
            ws.cell(row=row_num, column=4).value = row['tidak_melanggar']

            col = 5
            for pl in row['dt_pelanggaran']:
                ws.cell(row=row_num, column=col).value = pl['total']
                col += 1
            for sn in row['dt_penindakan']:
                ws.cell(row=row_num, column=col).value = sn['total']
                col += 1
            for ss in row['dt_subsanksi']:
                ws.cell(row=row_num, column=col).value = ss['total']
                col += 1

            row_num += 1

        # ====== Styling Sederhana ======
        for r in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=col-1):
            for cell in r:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="BAD8B6", end_color="BAD8B6", fill_type="solid")

        for i in range(1, col):
            ws.column_dimensions[chr(64 + i)].width = 15

        # ====== Kirim Response ======
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rekap_Pengawasan_'+lokasi.nama.lower()+'_'+now().strftime("%Y%m%d%H%M%S")+'.xlsx"'
        wb.save(response)
        return response
    elif export_type == 'pdf':
        nama_korsatpel = '-'
        nip_korsatpel = '-'
        pangkat_korsatpel = '-'
        ttd_korsatpel = None
        if lokasi.korsatpel_id:
            korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
            base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
            fallback = None
            def url_or_fallback(fn):
                if not fn:
                    return fallback
                if not os.path.isfile(os.path.join(base_dir, fn)):
                    return fallback
                return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

            if korsatpel:
                nama_korsatpel = korsatpel.name
                nip_korsatpel = korsatpel.nip
                pangkat_korsatpel = korsatpel.pangkat
                ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)
        # ====== Buat Header Dinamis ======
        jumlah_pelanggaran = len(jenisPelanggaran())
        jumlah_sanksi = len(sanksi())
        jumlah_subsanksi = len(subSanksi())

        headers_row1 = [
            {"label": "WAKTU", "colspan": 1, "rowspan": 2},
            {"label": "JUMLAH KENDARAAN", "colspan": 3, "rowspan": 1},
            {"label": "PELANGGARAN", "colspan": jumlah_pelanggaran, "rowspan": 1},
            {"label": "PENINDAKAN", "colspan": jumlah_sanksi, "rowspan": 1},
            {"label": "PENINDAKAN LAINNYA", "colspan": jumlah_subsanksi, "rowspan": 1},
        ]

        headers_row2 = ["DIPERIKSA", "MELANGGAR", "TIDAK MELANGGAR"] \
                    + [pl['name'] for pl in jenisPelanggaran()] \
                    + [sn['name'] for sn in sanksi()] \
                    + [ss['name'] for ss in subSanksi()]

        # ====== Kirim data ke template ======
        context = {
            "lokasi": lokasi.nama.upper(),
            "headers_row1": headers_row1,
            "headers_row2": headers_row2,
            "periode_name": (
                bulan_indo[filter_month - 1].upper() + ' ' + str(filter_year)
                if filter_interval == 'MONTH'
                else str(filter_year)
            ),
            "korsatpel": {
                "nama": nama_korsatpel.upper(),
                "nip": nip_korsatpel.upper(),
                "pangkat": pangkat_korsatpel.upper(),
                "ttd": ttd_korsatpel,
            },
            "gambar": {
                "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
                "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
            },
            "data": data,  # sudah berisi baris TOTAL juga
        }
        html = render_to_string("exports/report_pengawasan.html", context)

        response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'inline; filename="Rekap_Pengawasan_'+lokasi.nama.lower()+'_'+now().strftime("%Y%m%d%H%M%S")+'.pdf"'
        response['Content-Disposition'] = 'atachment; filename="Rekap_Pengawasan_'+lokasi.nama.lower()+'_'+now().strftime("%Y%m%d%H%M%S")+'.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Terjadi kesalahan saat membuat PDF', status=500)
        return response