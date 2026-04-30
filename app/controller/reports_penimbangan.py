from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, DataPenimbangan, MasterPelanggaran, DataPelanggaran, MasterJenisPelanggaran, MasterKomoditi, DataKotaKab, MasterTimbangan, MasterRegu, MasterShift, MasterJenisKendaraan, UserSystem, DataSdm
from django.templatetags.static import static
from django.db.models import Q
from datetime import date
from .utils import json_response
import locale, os
from datetime import datetime, date, time
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import random
from django.conf import settings

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Data Penimbangan',
        'active_menu': 'laporan_penimbangan',
        'app_name': lokasi.nama,
    }
    return render(request, 'backend/reports_penimbangan.html', context)


def show(request):
    if 'is_widget' in request.GET:
        filter_shift = request.GET.get('filter_shift')
        filter_regu = request.GET.get('filter_regu')
        filter_timbangan = request.GET.get('filter_timbangan')
        filter_date = str(request.GET.get('filter_date') or '')
        search = request.GET.get('search')

        def parse_id_datetime(s: str, is_end=False):
            """Parse datetime dari format Indonesian DD/MM/YYYY tanpa timezone"""
            s = s.strip()
            for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    # Kalau hanya tanggal
                    if fmt == "%d/%m/%Y":
                        if is_end:
                            dt = datetime.combine(dt.date(), time.max)
                        else:
                            dt = datetime.combine(dt.date(), time.min)
                    return dt
                except ValueError:
                    continue
            raise ValueError(f"Format tanggal tidak valid: {s}")

        # ===== parsing filter =====
        start_str, end_str = filter_date.split(" - ")

        start_datetime = parse_id_datetime(start_str, is_end=False)
        end_datetime = parse_id_datetime(end_str, is_end=True)

        # ===== Query =====
        base_filter = DataPenimbangan.objects.filter(
            tgl_penimbangan__range=(start_datetime, end_datetime),
            is_transaksi=1
        )
        if filter_shift and filter_shift != 'ALL':
            base_filter = base_filter.filter(
                shift_id = filter_shift
            )
        if filter_regu and filter_regu != 'ALL':
            base_filter = base_filter.filter(
                regu_id = filter_regu
            )
        if filter_timbangan and filter_timbangan != 'ALL':
            base_filter = base_filter.filter(
                timbangan_id = filter_timbangan
            )
        if search:
            base_filter = base_filter.filter(
                Q(no_kendaraan__icontains=search) |
                Q(no_uji__icontains=search) |
                Q(nama_pemilik__icontains=search)
            )
        kendaraan_timbang = base_filter.count()
        kendaraan_melanggar = base_filter.filter(is_melanggar='Y').count()
        kendaraan_tidak_melanggar = base_filter.filter(is_melanggar='N').count()

        if kendaraan_timbang > 0:
            persentase_kendaraan_melanggar = round((kendaraan_melanggar / kendaraan_timbang) * 100, 2)
            persentase_kendaraan_tidak_melanggar = round((kendaraan_tidak_melanggar / kendaraan_timbang) * 100, 2)
        else:
            persentase_kendaraan_melanggar = 0.00
            persentase_kendaraan_tidak_melanggar = 0.00
            
        output = {
            'kendaraan_timbang': f"{kendaraan_timbang:,}".replace(",", "."),
            'kendaraan_melanggar': f"{kendaraan_melanggar:,}".replace(",", "."),
            'kendaraan_tidak_melanggar': f"{kendaraan_tidak_melanggar:,}".replace(",", "."),
            'persentase_kendaraan_melanggar': persentase_kendaraan_melanggar,
            'persentase_kendaraan_tidak_melanggar': persentase_kendaraan_tidak_melanggar,
        }
        return json_response(status=True, message='success', data=output)
    else:
        try:
            # Params datatables
            draw = int(request.GET.get('draw', 0))
            start = int(request.GET.get('start', 0))
            length = int(request.GET.get('length', 10))
            search = request.GET.get('search[value]', '').strip()

            filter_shift = request.GET.get('filter_shift')
            filter_regu = request.GET.get('filter_regu')
            filter_timbangan = request.GET.get('filter_timbangan')
            filter_date = str(request.GET.get('filter_date') or '')
            print("="*40)
            print("FILTER PARAMETER:")
            print(f"Shift     : {filter_shift or '-'}")
            print(f"Regu      : {filter_regu or '-'}")
            print(f"Timbangan : {filter_timbangan or '-'}")
            print(f"Tanggal   : {filter_date or '-'}")
            print("="*40)

            def parse_id_datetime(s: str) -> datetime:
                """Parse datetime dari format Indonesian DD/MM/YYYY tanpa timezone"""
                s = (s or "").strip()
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
                    try:
                        dt = datetime.strptime(s, fmt)
                        # Kalau cuma tanggal, set jam default
                        if fmt == "%d/%m/%Y":
                            dt = datetime.combine(dt.date(), time.min)
                        return dt
                    except ValueError:
                        continue
                raise ValueError(f"Format tanggal tidak valid: {s}")
            
            start_str, end_str = [x.strip() for x in filter_date.split(" - ")]

            start_datetime = parse_id_datetime(start_str)
            end_datetime = parse_id_datetime(end_str)

            # Set end time ke 23:59:59.999999 untuk include full day
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59, microsecond=999999)

            query = DataPenimbangan.objects.order_by("-tgl_penimbangan").filter(
                tgl_penimbangan__range=(start_datetime, end_datetime)
            )

            if filter_shift and filter_shift != 'ALL':
                query = query.filter(
                    shift_id = filter_shift
                )
            if filter_regu and filter_regu != 'ALL':
                query = query.filter(
                    regu_id = filter_regu
                )
            if filter_timbangan and filter_timbangan != 'ALL':
                query = query.filter(
                    timbangan_id = filter_timbangan
                )
            if search:
                query = query.filter(
                    Q(no_kendaraan__icontains=search) |
                    Q(no_uji__icontains=search) |
                    Q(nama_pemilik__icontains=search)
                )
            getRow = query.all()
            total_record = query.count()
            paginated_record = getRow[start:start+length]
            data = []

            for index, penimbangan in enumerate(paginated_record, start=start + 1):
                melanggarCustom = '<span class="badge badge-success fs-7">TIDAK MELANGGAR</span>' if penimbangan.is_melanggar == 'N' else '<span class="badge badge-danger fs-7">MELANGGAR</span>'

                statusPengiriman = '<span class="badge badge-success fs-7 text-light">TERKIRIM</span>'
                if penimbangan.is_send_to_pusat == 'N':
                    statusPengiriman = f"""
                        <button type="button" 
                                class="btn btn-sm btn-warning ms-1" 
                                data-bs-toggle="tooltip" 
                                title="Kirim Data!" 
                                onclick="sendDataPenimbangan('{penimbangan.id}');">Kirim
                        </button>
                    """
                action = '-'
                if penimbangan.is_send_to_pusat == 'N':
                    action = f'''
                        <button type="button" class="btn btn-sm btn-success mb-1 ms-1" data-bs-toggle="tooltip" title="Kirim penimbangan!" onclick="sendDataPenimbangan('{penimbangan.id}');"><i class="ki-outline ki-send"></i>Kirim</button>
                    '''
                # Foto Penimbangan Depan
                foto_depan = static('dist/img/default-img.png')
                if penimbangan.foto_depan:
                    if os.path.exists(os.path.join('static', 'dist', 'img', 'penimbangan', penimbangan.foto_depan)):
                        foto_depan = static('dist/img/penimbangan/') + penimbangan.foto_depan

                # Foto Penimbangan Belakang
                foto_belakang = static('dist/img/default-img.png')
                if penimbangan.foto_belakang:
                    if os.path.exists(os.path.join('static', 'dist', 'img', 'penimbangan', penimbangan.foto_belakang)):
                        foto_belakang = static('dist/img/penimbangan/') + penimbangan.foto_belakang
                foto_depan = f'''
                    <a class="image-popup" href="{foto_depan}" data-bs-toggle="tooltip" title="Klik untuk melihat gambar!">
                        <img src="{foto_depan}" class="rounded" alt="foto-penimbangan1" title="FOTO PENIMBANGAN - 1" width="72px">
                    </a>
                '''
                foto_belakang = f'''
                    <a class="image-popup" href="{foto_belakang}" data-bs-toggle="tooltip" title="Klik untuk melihat gambar!">
                        <img src="{foto_belakang}" class="rounded" alt="foto-penimbangan1" title="FOTO PENIMBANGAN - 2" width="72px">
                    </a>
                '''
                
                pelanggaran = 'TIDAK MELANGGAR'
                if penimbangan.is_melanggar == 'Y':
                    pelanggaran_qs = DataPelanggaran.objects.filter(kode_trx=penimbangan.kode_trx)
                    if pelanggaran_qs.exists():
                        pelanggaran = ''
                        for pel in pelanggaran_qs:
                            try:
                                jenis_pelanggaran = MasterJenisPelanggaran.objects.get(
                                    id=int(pel.jenis_pelanggaran_id)
                                )
                                pelanggaran += '<span class="badge badge-light me-1 text-dark">' + jenis_pelanggaran.nama + '</span>'
                            except MasterJenisPelanggaran.DoesNotExist:
                                continue
                # ==============================
                # KOMODITI, ASAL, TUJUAN, PENGADILAN
                # ==============================
                komoditi = MasterKomoditi.objects.filter(id=penimbangan.komoditi_id).select_related('kategori_komoditi').first()
                asal_kota = DataKotaKab.objects.filter(id=penimbangan.asal_kota_id).first()
                tujuan_kota = DataKotaKab.objects.filter(id=penimbangan.tujuan_kota_id).first()
                timbangan = MasterTimbangan.objects.filter(id=penimbangan.timbangan_id).first()
                regu = MasterRegu.objects.filter(id=penimbangan.regu_id).first()
                shift = MasterShift.objects.filter(id=penimbangan.shift_id).first()
                jenis_kendaraan = MasterJenisKendaraan.objects.filter(id=penimbangan.jenis_kendaraan_id).first()
                operator = UserSystem.objects.filter(id=penimbangan.petugas_id).first()
                
                data.append({
                    'status': statusPengiriman,
                    'melanggar': melanggarCustom,
                    'action': action,
                    'waktu': penimbangan.tgl_penimbangan.strftime('%d/%m/%Y %H:%M'),
                    'timbangan': timbangan.nama if timbangan else '-',
                    'foto_depan': foto_depan,
                    'foto_belakang': foto_belakang,
                    'no_kendaraan': penimbangan.no_kendaraan,
                    'no_uji': penimbangan.no_uji,
                    'masa_berlaku_uji': penimbangan.tgl_masa_berlaku,
                    'jenis_kendaraan': jenis_kendaraan.nama if jenis_kendaraan else '-',
                    'sumbu': penimbangan.sumbu,
                    'jbi': penimbangan.jbi_uji,
                    'berat_timbang': penimbangan.berat_timbang,
                    'kelebihan_berat': penimbangan.kelebihan_berat,
                    'prosen_lebih': penimbangan.prosen_lebih,
                    'toleransi': str(penimbangan.toleransi_komoditi),
                    
                    'panjang_utama': int(penimbangan.panjang_utama),
                    'lebar_utama': int(penimbangan.lebar_utama),
                    'tinggi_utama': int(penimbangan.tinggi_utama),
                    
                    'panjang_ukur': int(penimbangan.panjang_ukur),
                    'lebar_ukur': int(penimbangan.lebar_ukur),
                    'tinggi_ukur': int(penimbangan.tinggi_ukur),

                    'panjang_lebih': int(penimbangan.panjang_lebih),
                    'lebar_lebih': int(penimbangan.lebar_lebih),
                    'tinggi_lebih': int(penimbangan.tinggi_lebih),
                    
                    'komoditi': komoditi.nama if komoditi else '-',
                    'asal_tujuan': (asal_kota.nama if asal_kota else '-') + ' - ' + (tujuan_kota.nama if tujuan_kota else '-'),
                    'pelanggaran': pelanggaran,
                    'regu': regu.nama if regu else '-',
                    'shift': shift.nama if shift else '-',
                    'operator': operator.nama if operator else '-',
                })


            return JsonResponse({
                'draw': draw,
                'recordsTotal': total_record,
                'recordsFiltered': getRow.count(),
                'data': data,
            })
        except Exception as e:
            return json_response(status=False, message=str(e), code=401) 
    
def export(request):
    export_type = request.GET.get('export_type')

    filter_date = request.GET.get('filter_date', '')
    search = request.GET.get('search[value]', '')
    filter_shift    = request.GET.get('filter_shift')
    filter_regu     = request.GET.get('filter_regu')
    filter_timbangan = request.GET.get('filter_timbangan')

    lokasi = LokasiUppkb.objects.first()


    def parse_id_datetime(s: str, is_end: bool = False) -> datetime:
        s = s.strip()
        for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                # kalau hanya tanggal, set jam min/max
                if fmt == "%d/%m/%Y":
                    dt = datetime.combine(dt.date(), time.max if is_end else time.min)
                return dt
            except ValueError:
                continue
        raise ValueError(f"Format tanggal tidak valid: {s}")

    start_str, end_str = filter_date.split(" - ")

    start_naive = parse_id_datetime(start_str, is_end=False)
    end_naive   = parse_id_datetime(end_str, is_end=True)

    start_dt = datetime.combine(start_naive, time.min)
    end_dt = datetime.combine(end_naive, time.max)
    # ============================
    # QUERY DASAR
    # ============================
    query = DataPenimbangan.objects.filter(
        tgl_penimbangan__range=(start_dt, end_dt)
    ).order_by('-tgl_penimbangan')

    if filter_shift and filter_shift != "ALL":
        query = query.filter(shift_id=filter_shift)
    if filter_regu and filter_regu != "ALL":
        query = query.filter(regu_id=filter_regu)
    if filter_timbangan and filter_timbangan != "ALL":
        query = query.filter(timbangan_id=filter_timbangan)
        
    # if not filter_shift:
    #     return json_response(status=False, message="Komoditi tidak ditemukan", code=400)

    if search:
        query = query.filter(
            Q(no_kendaraan__icontains=search) |
            Q(no_uji__icontains=search) |
            Q(nama_pemilik__icontains=search)
        )

    # ============================
    # SERIALIZE DATA UNTUK EXCEL/PDF
    # ============================
    data = []
    jumlah_penimbangan = 0
    jumlah_melanggar = 0
    jumlah_tidak_melanggar = 0
    for p in query:
        melanggar = "TIDAK MELANGGAR" if p.is_melanggar == "N" else "MELANGGAR"
        
        tgl_local = p.tgl_penimbangan.strftime('%d-%m-%Y %H:%M:%S')

        # Ambil Pelanggaran
        pelanggaran = "TIDAK MELANGGAR"
        jumlah_penimbangan += 1
        if p.is_melanggar == "Y":
            jumlah_melanggar += 1
            pel_qs = DataPelanggaran.objects.filter(kode_trx=p.kode_trx)
            if pel_qs.exists():
                pelanggaran = ", ".join([
                    MasterJenisPelanggaran.objects.get(id=int(i.jenis_pelanggaran_id)).nama
                    for i in pel_qs
                    if MasterJenisPelanggaran.objects.filter(id=int(i.jenis_pelanggaran_id)).exists()
                ])
        else:
            jumlah_tidak_melanggar += 1
        operator = UserSystem.objects.filter(id=p.petugas_id).first()
        data.append({
            "waktu": tgl_local,
            "uppkb": lokasi.nama.upper(),
            "no_kendaraan": p.no_kendaraan,
            "no_uji": p.no_uji,
            "pemilik_komoditi": p.pemilik_komoditi,
            "alamat_pemilik_komoditi": p.alamat_pemilik_komoditi,
            "jenis_kendaraan": MasterJenisKendaraan.objects.get(id=p.jenis_kendaraan_id).nama if p.jenis_kendaraan_id else "-",
            "sumbu": p.sumbu,
            "jbi": int(p.jbi_uji),
            "berat_timbang": p.berat_timbang,
            "kelebihan_berat": int(p.kelebihan_berat),
            "prosen_lebih": p.prosen_lebih,
            "toleransi": str(p.toleransi_komoditi),

            'panjang_utama': int(p.panjang_utama),
            'lebar_utama': int(p.lebar_utama),
            'tinggi_utama': int(p.tinggi_utama),
            
            'panjang_ukur': int(p.panjang_ukur),
            'lebar_ukur': int(p.lebar_ukur),
            'tinggi_ukur': int(p.tinggi_ukur),

            'panjang_lebih': int(p.panjang_lebih),
            'lebar_lebih': int(p.lebar_lebih),
            'tinggi_lebih': int(p.tinggi_lebih),
            
            "asal": DataKotaKab.objects.get(id=p.asal_kota_id).nama if p.asal_kota_id else "-",
            "tujuan": DataKotaKab.objects.get(id=p.tujuan_kota_id).nama if p.tujuan_kota_id else "-",
            "komoditi": MasterKomoditi.objects.get(id=p.komoditi_id).nama if p.komoditi_id else "-",
            "pelanggaran": pelanggaran,
            "shift": MasterShift.objects.get(id=p.shift_id).nama if p.shift_id else "-",
            "regu": MasterRegu.objects.get(id=p.regu_id).nama if p.regu_id else "-",
            'operator' : operator.nama if operator else '-',
        })
        print(operator.nama if operator else '-')

    # ============================
    # EXPORT EXCEL
    # ============================
    if export_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Penimbangan"

        HEADER_MAP = {
            "waktu": "WAKTU PENIMBANGAN",
            "uppkb": "UPPKB",
            "no_kendaraan": "NO KENDARAAN",
            "no_uji": "NO UJI",
            "pemilik_komoditi": "NAMA PEMILIK",
            "alamat_pemilik_komoditi": "ALAMAT PEMILIK",
            "jenis_kendaraan": "JENIS KENDARAAN",
            "sumbu": "SUMBU",
            "jbi": "JBI(Kg)",
            "berat_timbang": "BERAT TIMBANG(Kg)",
            "kelebihan_berat": "BERAT LEBIH(Kg)",
            "prosen_lebih": "PERSEN LEBIH(%)",
            "toleransi": "TOLERANSI(%)",

            'panjang_utama': 'PANJANG UTAMA',
            'lebar_utama': 'LEBAR UTAMA',
            'tinggi_utama': 'TINGGI UTAMA',
            
            'panjang_ukur': 'PANJANG UKUT',
            'lebar_ukur': 'LEBAR UKUR',
            'tinggi_ukur': 'TINGGI UKUR',

            'panjang_lebih': 'PANJANG LEBIH',
            'lebar_lebih': 'LEBAR LEBIH',
            'tinggi_lebih': 'TINGGI LEBIH',
            
            "asal": "ASAL",
            "tujuan": "TUJUAN",
            "komoditi": "KOMODITI",
            "pelanggaran": "PELANGGARAN",
            "shift": "SHIFT",
            "regu": "REGU",
            "operator": "OPERATOR",
        }
        headers = [HEADER_MAP.get(k, k.upper()) for k in data[0].keys()] if data else []
        ws.append(headers)

        # ===== Styling =====
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Border tebal untuk header (medium)
        header_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Border tipis untuk body data
        body_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== SET HEADER =====
        for h in ws[1]:
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = orange_fill
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = header_border

        # ===== SET BODY =====
        # Append data rows
        for row in data:
            ws.append(list(row.values()))

        # Apply border ke setiap cell di body
        row_start = 2
        row_end = ws.max_row
        col_start = 1
        col_end = ws.max_column

        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = body_border

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = now().strftime("%Y_%m_%d_%H_%M_%S_%f")
        random_code = random.randint(10000000, 99999999)
        filename = f'penimbangan_{timestamp}_{random_code}.xlsx'
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    # ============================
    # EXPORT PDF
    # ============================
    nama_korsatpel = '-'
    nip_korsatpel = '-'
    pangkat_korsatpel = '-'
    ttd_korsatpel = None
    if lokasi.korsatpel_id:
        korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
        base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
        fallback = None
        def url_or_fallback(fn):
            if not fn:
                return fallback
            if not os.path.isfile(os.path.join(base_dir, fn)):
                return fallback
            return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

        if korsatpel:
            nama_korsatpel = korsatpel.name
            nip_korsatpel = korsatpel.nip
            pangkat_korsatpel = korsatpel.pangkat
            ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)
            
    context = {
        "lokasi": lokasi.nama.upper(),
        "periode_name": filter_date,
        "date_now": now().strftime("%d-%m-%Y"),
        "shift": MasterShift.objects.get(id=filter_shift).nama if filter_shift else "SEMUA",
        "regu": MasterRegu.objects.get(id=filter_regu).nama if filter_regu else "SEMUA",
        "korsatpel": {
            "nama": nama_korsatpel.upper(),
            "nip": nip_korsatpel.upper(),
            "pangkat": pangkat_korsatpel.upper(),
            "ttd": ttd_korsatpel,
        },
        "jumlah": {
            "penimbangan" : jumlah_penimbangan,
            "melanggar" : jumlah_melanggar,
            "tidak_melanggar" : jumlah_tidak_melanggar,
        },
        "gambar": {
            "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
            "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
        },
        "data": data,
    }

    html = render_to_string("exports/report_penimbangan.html", context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        # f'inline; filename="data_penimbangan_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
        f'atachment; filename="data_penimbangan_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)

    return response