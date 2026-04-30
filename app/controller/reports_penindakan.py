from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from app.models import LokasiUppkb, DataPenimbangan, MasterPelanggaran, DataPelanggaran, MasterJenisPelanggaran, MasterKomoditi, DataKotaKab, MasterTimbangan, MasterRegu, MasterShift, MasterJenisKendaraan, UserSystem, DataPenindakan, Pengadilan, Sanksi, Sitaan, SubSanksi, DataSdm
from django.templatetags.static import static
from django.db.models import Q
from datetime import date, datetime, time
from .utils import json_response
import locale, os
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import random
from django.conf import settings

def index(request):
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
    lokasi = LokasiUppkb.objects.first()
    context = {
        'title': 'Data Penindakan',
        'active_menu': 'laporan_penindakan',
        'app_name': lokasi.nama,
    }
    return render(request, 'backend/reports_penindakan.html', context)


def show(request):
    try:
        # Params datatables
        draw = int(request.GET.get('draw', 0))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search = request.GET.get('search[value]', '').strip()

        filter_shift = request.GET.get('filter_shift')
        filter_regu = request.GET.get('filter_regu')
        filter_sanksi = request.GET.get('filter_sanksi')
        filter_date = str(request.GET.get('filter_date') or '')
        
        print("="*40)
        print("FILTER PARAMETER:")
        print(f"Shift     : {filter_shift or '-'}")
        print(f"Regu      : {filter_regu or '-'}")
        print(f"Sanksi    : {filter_sanksi or '-'}")
        print(f"Tanggal   : {filter_date or '-'}")
        print("="*40)
        
        def parse_id_datetime(s: str, is_end: bool = False) -> datetime:
            """Parse datetime dari format Indonesian DD/MM/YYYY tanpa timezone"""
            s = s.strip()
            for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    # kalau hanya tanggal → set jam default
                    if fmt == "%d/%m/%Y":
                        dt = datetime.combine(dt.date(), time.max if is_end else time.min)
                    return dt
                except ValueError:
                    continue
            raise ValueError(f"Format tanggal tidak valid: {s}")

        # ============================
        # FILTER TANGGAL + JAM
        # ============================
        start_str, end_str = filter_date.split(" - ")

        start_datetime = parse_id_datetime(start_str, is_end=False)
        end_datetime   = parse_id_datetime(end_str, is_end=True)

        # ============================
        # QUERY
        # ============================
        query = DataPenindakan.objects.order_by('-tgl_penindakan').filter(
            tgl_penindakan__range=(start_datetime, end_datetime)
        )
        if filter_shift and filter_shift != 'ALL':
            query = query.filter(
                shift_id = filter_shift
            )
        if filter_regu and filter_regu != 'ALL':
            query = query.filter(
                regu_id = filter_regu
            )
        if filter_sanksi and filter_sanksi != 'ALL':
            query = query.filter(
                sanksi_id = filter_sanksi
            )
        if search:
            query = query.filter(
                Q(nama_pengemudi__icontains=search) |
                Q(alamat_pengemudi__icontains=search)
            )
        getRow = query.all()
        total_record = query.count()
        paginated_record = getRow[start:start+length]
        data = []
        # return True
        for index, penindakan in enumerate(paginated_record, start=start + 1):
            statusPengiriman = '<span class="badge badge-success fs-7 text-light">TERKIRIM</span>'
            if penindakan.is_send_to_pusat == 'N':
                statusPengiriman = f"""
                    <button type="button" 
                            class="btn btn-sm btn-warning ms-1" 
                            data-bs-toggle="tooltip" 
                            title="Kirim Data!" 
                            onclick="sendDataPenindakan('{penindakan.id}');">Kirim
                    </button>
                """

            penimbangan = DataPenimbangan.objects.filter(kode_trx=penindakan.kode_trx).first()
            action = f"""
                <button type="button" 
                        class="btn btn-sm btn-dark ms-1" 
                        data-bs-toggle="tooltip" 
                        title="Cetak Surat Peringatan!" 
                        onclick="_printSuratPeringatan('Y', '{penimbangan.id}');">
                    <i class="ki-solid ki-printer fs-6"></i>Peringatan
                </button>
            """
            if penindakan.sanksi_id == 11:
                action = f"""
                    <button type="button" 
                            class="btn btn-sm btn-dark ms-1" 
                            data-bs-toggle="tooltip" 
                            title="Cetak Blanko Tilang!" 
                            onclick="_printSuratTilang('Y', '{penimbangan.id}');">
                        <i class="ki-solid ki-printer fs-6"></i>Blanko
                    </button>
                """        
            
            pelanggaran = 'TIDAK MELANGGAR'
            if penimbangan.is_melanggar == 'Y':
                pelanggaran_qs = DataPelanggaran.objects.filter(kode_trx=penimbangan.kode_trx)
                if pelanggaran_qs.exists():
                    pelanggaran = ''
                    for pel in pelanggaran_qs:
                        try:
                            jenis_pelanggaran = MasterJenisPelanggaran.objects.get(
                                id=int(pel.jenis_pelanggaran_id)
                            )
                            pelanggaran += '<span class="badge badge-light me-1 text-dark">' + jenis_pelanggaran.nama + '</span>'
                        except MasterJenisPelanggaran.DoesNotExist:
                            continue

            sanksi = Sanksi.objects.filter(id=penindakan.sanksi_id).first()
            asal_kota = DataKotaKab.objects.filter(id=penimbangan.asal_kota_id).first()
            tujuan_kota = DataKotaKab.objects.filter(id=penimbangan.tujuan_kota_id).first()
            pengadilan = Pengadilan.objects.filter(id=penindakan.pengadilan_id).first()

            sitaan_name = '-'
            if penindakan.sanksi_id == 11:
                if penindakan.sitaan:
                    sitaan_ids = [int(x) for x in penindakan.sitaan.split(',') if x.strip().isdigit()]
                    sitaan_qs = Sitaan.objects.filter(id__in=sitaan_ids)
                    sitaan_name = ", ".join(s.keterangan for s in sitaan_qs)

            sanksi_tambahan = '-'
            if penindakan.sanksi_id == 11:
                if penindakan.sanksi_tambahan:
                    sanksi_tambahan_ids = [int(x) for x in penindakan.sanksi_tambahan.split(',') if x.strip().isdigit()]
                    sanksi_tambahan_qs = SubSanksi.objects.filter(id__in=sanksi_tambahan_ids)
                    sanksi_tambahan = ", ".join(s.keterangan for s in sanksi_tambahan_qs)
            
            data.append({
                'sink': statusPengiriman,
                'action': action,
                'waktu': penindakan.tgl_penindakan.strftime('%d/%m/%Y %H:%M'),
                'no_kendaraan': penimbangan.no_kendaraan,
                'nama_pengemudi': penindakan.nama_pengemudi,
                'alamat_pengemudi': penindakan.alamat_pengemudi,
                'asal_tujuan': asal_kota.nama + '-' + tujuan_kota.nama,
                'no_uji': penimbangan.no_uji,
                'pelanggaran': pelanggaran,
                'sanksi': sanksi.nama,
                'sitaan': sitaan_name,
                'sanksi_tambahan': sanksi_tambahan,
                'tgl_sidang': penindakan.tgl_sidang,
                'jam_sidang': penindakan.jam_sidang,
                'pengadilan': pengadilan.nama,
                'no_skep': penindakan.nip_ppns,
                'ppns': penindakan.nama_ppns,
                'keterangan': penindakan.keterangan_tindakan,
            })


        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_record,
            'recordsFiltered': getRow.count(),
            'data': data,
        })
    except Exception as e:
        return json_response(status=False, message=str(e), code=401) 
    
def export(request):
    export_type = request.GET.get('export_type')

    search = request.GET.get('search[value]', '')
    filter_shift = request.GET.get('filter_shift')
    filter_regu = request.GET.get('filter_regu')
    filter_sanksi = request.GET.get('filter_sanksi')
    filter_date = str(request.GET.get('filter_date') or '')

    lokasi = LokasiUppkb.objects.first()

    def parse_id_datetime(s: str, is_end: bool = False) -> datetime:
        s = s.strip()
        for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                # kalau hanya tanggal → set jam default
                if fmt == "%d/%m/%Y":
                    dt = datetime.combine(dt.date(), time.max if is_end else time.min)
                return dt
            except ValueError:
                continue
        raise ValueError(f"Format tanggal tidak valid: {s}")

    # ============================
    # FILTER TANGGAL + JAM
    # ============================
    start_str, end_str = filter_date.split(" - ")

    start_naive = parse_id_datetime(start_str, is_end=False)
    end_naive   = parse_id_datetime(end_str, is_end=True)

    start_datetime = parse_id_datetime(start_str, is_end=False)
    end_datetime   = parse_id_datetime(end_str, is_end=True)

    # ============================
    # QUERY
    # ============================
    query = DataPenindakan.objects.order_by('-tgl_penindakan').filter(
        tgl_penindakan__range=(start_datetime, end_datetime)
    )
    if filter_shift and filter_shift != 'ALL':
        query = query.filter(
            shift_id = filter_shift
        )
    if filter_regu and filter_regu != 'ALL':
        query = query.filter(
            regu_id = filter_regu
        )
    if filter_sanksi and filter_sanksi != 'ALL':
        query = query.filter(
            sanksi_id = filter_sanksi
        )
    if search:
        query = query.filter(
            Q(nama_pengemudi__icontains=search) |
            Q(alamat_pengemudi__icontains=search)
        )

    # ============================
    # SERIALIZE DATA UNTUK EXCEL/PDF
    # ============================
    data = []
    for penindakan in query:
        penimbangan = DataPenimbangan.objects.filter(kode_trx=penindakan.kode_trx).first()         
        pelanggaran = 'TIDAK MELANGGAR'
        if penimbangan.is_melanggar == 'Y':
            pelanggaran_qs = DataPelanggaran.objects.filter(kode_trx=penimbangan.kode_trx)
            if pelanggaran_qs.exists():
                pelanggaran = ''
                for pel in pelanggaran_qs:
                    try:
                        jenis_pelanggaran = MasterJenisPelanggaran.objects.get(
                            id=int(pel.jenis_pelanggaran_id)
                        )
                        pelanggaran += jenis_pelanggaran.nama + ', '
                    except MasterJenisPelanggaran.DoesNotExist:
                        continue

        sanksi = Sanksi.objects.filter(id=penindakan.sanksi_id).first()
        asal_kota = DataKotaKab.objects.filter(id=penimbangan.asal_kota_id).first()
        tujuan_kota = DataKotaKab.objects.filter(id=penimbangan.tujuan_kota_id).first()
        pengadilan = Pengadilan.objects.filter(id=penindakan.pengadilan_id).first()

        sitaan_name = '-'
        if penindakan.sanksi_id == 11:
            if penindakan.sitaan:
                sitaan_ids = [int(x) for x in penindakan.sitaan.split(',') if x.strip().isdigit()]
                sitaan_qs = Sitaan.objects.filter(id__in=sitaan_ids)
                sitaan_name = ", ".join(s.keterangan for s in sitaan_qs)

        sanksi_tambahan = '-'
        if penindakan.sanksi_id == 11:
            if penindakan.sanksi_tambahan:
                sanksi_tambahan_ids = [int(x) for x in penindakan.sanksi_tambahan.split(',') if x.strip().isdigit()]
                sanksi_tambahan_qs = SubSanksi.objects.filter(id__in=sanksi_tambahan_ids)
                sanksi_tambahan = ", ".join(s.keterangan for s in sanksi_tambahan_qs)

        data.append({
            'waktu': penindakan.tgl_penindakan.strftime('%d/%m/%Y %H:%M'),
            'no_kendaraan': penimbangan.no_kendaraan,
            'nama_pengemudi': penindakan.nama_pengemudi,
            'alamat_pengemudi': penindakan.alamat_pengemudi,
            'asal_tujuan': asal_kota.nama + '-' + tujuan_kota.nama,
            'no_uji': penimbangan.no_uji,
            'pelanggaran': pelanggaran,
            'sanksi': sanksi.nama,
            'sitaan': sitaan_name,
            'sanksi_tambahan': sanksi_tambahan,
            'tgl_sidang': penindakan.tgl_sidang,
            'jam_sidang': penindakan.jam_sidang,
            'pengadilan': pengadilan.nama,
            'no_skep': penindakan.nip_ppns,
            'ppns': penindakan.nama_ppns,
            'keterangan': penindakan.keterangan_tindakan,
        })

    # ============================
    # EXPORT EXCEL
    # ============================
    if export_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Penindakan"

        HEADER_MAP = {
            "waktu": "WAKTU PENINDAKAN",
            "no_kendaraan": "NO KENDARAAN",
            "nama_pengemudi": "NAMA PENGEMUDI",
            "alamat_pengemudi": "ALAMAT PENGEMUDI",
            "asal_tujuan": "ASAL - TUJUAN",
            "no_uji": "NO UJI",
            "pelanggaran": "PELANGGARAN",
            "sanksi": "SANKSI",
            "sitaan": "SITAAN",
            "sanksi_tambahan": "SANKSI TAMBAHAN",
            "tgl_sidang": "TGL SIDANG",
            "jam_sidang": "JAM SIDANG",
            "pengadilan": "PENGADILAN",
            "no_skep": "NOMOR SKEP",
            "ppns": "PPNS",
            "keterangan": "KETERANGAN",
        }
        headers = [HEADER_MAP.get(k, k.upper()) for k in data[0].keys()] if data else []
        ws.append(headers)

        # ===== Styling =====
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Border tebal untuk header (medium)
        header_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Border tipis untuk body data
        body_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== SET HEADER =====
        for h in ws[1]:
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = orange_fill
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = header_border

        # ===== SET BODY =====
        # Append data rows
        for row in data:
            ws.append(list(row.values()))

        # Apply border ke setiap cell di body
        row_start = 2
        row_end = ws.max_row
        col_start = 1
        col_end = ws.max_column

        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = body_border

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = now().strftime("%Y_%m_%d_%H_%M_%S_%f")
        random_code = random.randint(10000000, 99999999)
        filename = f'penindakan_{timestamp}_{random_code}.xlsx'
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    # ============================
    # EXPORT PDF
    # ============================
    nama_korsatpel = '-'
    nip_korsatpel = '-'
    pangkat_korsatpel = '-'
    ttd_korsatpel = None
    if lokasi.korsatpel_id:
        korsatpel = DataSdm.objects.filter(id=lokasi.korsatpel_id).first()
        base_dir = os.path.join(settings.BASE_DIR, 'static', 'dist', 'img', 'tanda-tangan')
        fallback = None
        def url_or_fallback(fn):
            if not fn:
                return fallback
            if not os.path.isfile(os.path.join(base_dir, fn)):
                return fallback
            return request.build_absolute_uri(static(f'dist/img/tanda-tangan/{fn}'))

        if korsatpel:
            nama_korsatpel = korsatpel.name
            nip_korsatpel = korsatpel.nip
            pangkat_korsatpel = korsatpel.pangkat
            ttd_korsatpel = url_or_fallback(korsatpel.ttd_korsatpel)
    context = {
        "lokasi": lokasi.nama.upper(),
        "periode_name": filter_date,
        "date_now": now().strftime("%d-%m-%Y"),
        "shift": MasterShift.objects.get(id=filter_shift).nama if filter_shift else "SEMUA",
        "regu": MasterRegu.objects.get(id=filter_regu).nama if filter_regu else "SEMUA",
        "korsatpel": {
            "nama": nama_korsatpel.upper(),
            "nip": nip_korsatpel.upper(),
            "pangkat": pangkat_korsatpel.upper(),
            "ttd": ttd_korsatpel,
        },
        "gambar": {
            "logo1": request.build_absolute_uri(static("dist/img/logo_kemenhub.png")),
            "logo2": request.build_absolute_uri(static("dist/img/logo_hubdat.png")),
        },
        "data": data,
    }

    html = render_to_string("exports/report_penindakan.html", context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'atachment; filename="data_penindakan_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
        # f'inline; filename="data_penindakan_{lokasi.nama.lower()}_{now().strftime("%Y%m%d%H%M%S")}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)

    return response